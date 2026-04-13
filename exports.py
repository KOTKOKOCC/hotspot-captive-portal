from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from db import fetch_all
from ui import format_dt, humanize_details
from labels import (
    COLUMN_LABELS,
    EVENT_LABELS,
    RESULT_LABELS,
    AUTH_METHOD_LABELS,
    STATUS_LABELS,
    TERMINATE_CAUSE_LABELS,
)



def rows_to_csv_bytes(rows, columns):
    sio = StringIO()
    writer = csv.writer(sio, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([row[col] if row[col] is not None else "" for col in columns])
    return sio.getvalue().encode("utf-8-sig")


def rows_to_xlsx_bytes(rows, columns, sheet_name="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    datetime_columns = {
        "created_at", "updated_at", "first_verified_at",
        "started_at", "ended_at", "expires_at",
        "event_time", "last_seen_at", "last_auth_at"
    }

    border_side = Side(style="medium", color="B8C4D6")
    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")

    # заголовки
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=COLUMN_LABELS.get(col_name, col_name))
        cell.font = Font(bold=True, size=12)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # данные
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row[col_name] if row[col_name] is not None else ""

            if col_name in datetime_columns:
                value = format_dt(value)
            elif col_name == "event_type":
                value = EVENT_LABELS.get(str(value), str(value))
            elif col_name == "result":
                value = RESULT_LABELS.get(str(value), str(value))
            elif col_name == "details":
                value = humanize_details(str(value))
            elif col_name == "auth_method":
                value = AUTH_METHOD_LABELS.get(str(value), str(value))
            elif col_name == "status":
                value = STATUS_LABELS.get(str(value).lower(), str(value))
            elif col_name == "terminate_cause":
                value = TERMINATE_CAUSE_LABELS.get(str(value), str(value))

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=11)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # заморозка шапки
    ws.freeze_panes = "A2"

    # автофильтр
    ws.auto_filter.ref = ws.dimensions

    # высота строки заголовка
    ws.row_dimensions[1].height = 36

    # автоширина колонок
    for col_idx, col_name in enumerate(columns, start=1):
        header_text = str(COLUMN_LABELS.get(col_name, col_name))
        max_len = len(header_text)

        for row_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))

        width = min(max(max_len + 4, 14), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def build_export_zip(date_from: str | None = None, date_to: str | None = None):
    def date_filter(col):
        if date_from and date_to:
            return f" WHERE date({col}) BETWEEN date('{date_from}') AND date('{date_to}') "
        if date_from:
            return f" WHERE date({col}) >= date('{date_from}') "
        if date_to:
            return f" WHERE date({col}) <= date('{date_to}') "
        return ""

    files = []

    guests = fetch_all("SELECT * FROM guests ORDER BY created_at DESC")
    files.append((
        "guests.csv",
        guests,
        ["id", "phone", "first_verified_at", "first_hotel", "auth_method", "status", "created_at", "updated_at"]
    ))

    sessions = fetch_all(f"SELECT * FROM guest_sessions {date_filter('started_at')} ORDER BY started_at DESC")
    files.append((
        "guest_sessions.csv",
        sessions,
        ["guest_id", "phone", "mac", "ip", "device_name", "started_at", "last_seen_at", "ended_at", "status", "terminate_cause", "acct_session_time", "hotel", "ssid", "vlan_id", "nas_id", "acct_session_id"]
    ))

    pending = fetch_all(f"SELECT * FROM pending_auth {date_filter('created_at')} ORDER BY created_at DESC")
    files.append((
        "pending_auth.csv",
        pending,
        ["id", "phone", "mac", "ip", "nas_id", "hotel", "ssid", "vlan_id", "created_at", "expires_at", "status"]
    ))

    calls = fetch_all(f"SELECT * FROM call_events {date_filter('created_at')} ORDER BY created_at DESC")
    files.append((
        "call_events.csv",
        calls,
        ["id", "phone", "callerid_raw", "source_ip", "created_at", "result"]
    ))

    audit = fetch_all(f"SELECT * FROM audit_log {date_filter('event_time')} ORDER BY event_time DESC")
    files.append((
        "audit_log.csv",
        audit,
        ["id", "phone", "mac", "ip", "nas_id", "hotel", "ssid", "vlan_id", "event_type", "event_time", "details"]
    ))

    networks = fetch_all("SELECT * FROM network_map ORDER BY vlan_id")
    files.append((
        "network_map.csv",
        networks,
        ["id", "hotel_name", "ssid_name", "vlan_id", "subnet_cidr", "mikrotik_interface", "hotspot_server", "is_active"]
    ))

    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, rows, cols in files:
            zf.writestr(filename, rows_to_csv_bytes(rows, cols))

    bio.seek(0)
    return bio


def build_single_xlsx(table_name: str, date_from: str | None = None, date_to: str | None = None):
    def date_filter(col):
        if date_from and date_to:
            return f" WHERE date({col}) BETWEEN date('{date_from}') AND date('{date_to}') "
        if date_from:
            return f" WHERE date({col}) >= date('{date_from}') "
        if date_to:
            return f" WHERE date({col}) <= date('{date_to}') "
        return ""

    mapping = {
        "guests": (
            "SELECT * FROM guests ORDER BY created_at DESC",
            ["id", "phone", "first_verified_at", "first_hotel", "auth_method", "status", "created_at", "updated_at"],
            "Guests"
        ),
        "sessions": (
            f"SELECT * FROM guest_sessions {date_filter('started_at')} ORDER BY started_at DESC",
            ["guest_id", "phone", "mac", "ip", "device_name", "started_at", "last_seen_at", "ended_at", "status", "terminate_cause", "acct_session_time", "hotel", "ssid", "vlan_id", "nas_id", "acct_session_id"],
            "Sessions"
        ),
        "pending": (
            f"SELECT * FROM pending_auth {date_filter('created_at')} ORDER BY created_at DESC",
            ["id", "phone", "mac", "ip", "nas_id", "hotel", "ssid", "vlan_id", "created_at", "expires_at", "status"],
            "Pending"
        ),
        "calls": (
            f"SELECT * FROM call_events {date_filter('created_at')} ORDER BY created_at DESC",
            ["id", "phone", "callerid_raw", "source_ip", "created_at", "result"],
            "Calls"
        ),
        "audit": (
            f"SELECT * FROM audit_log {date_filter('event_time')} ORDER BY event_time DESC",
            ["id", "phone", "mac", "ip", "nas_id", "hotel", "ssid", "vlan_id", "event_type", "event_time", "details"],
            "Audit"
        ),
        "networks": (
            "SELECT * FROM network_map ORDER BY vlan_id",
            ["id", "hotel_name", "ssid_name", "vlan_id", "subnet_cidr", "mikrotik_interface", "hotspot_server", "is_active"],
            "Networks"
        ),
    }

    if table_name not in mapping:
        raise HTTPException(status_code=404, detail="unknown export")

    query, columns, sheet_name = mapping[table_name]
    rows = fetch_all(query)
    return rows_to_xlsx_bytes(rows, columns, sheet_name)
